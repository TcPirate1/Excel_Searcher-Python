import openpyxl
import os
from path import RootPath

def find_cardName(currentSheet, Card_Name):
    for row in range(1, currentSheet.max_row + 1):
        for column in range(1,703): #columns
            cell1Column = num_hash(column)
            cell2Column = num_hash(column - 1)
            cell3Column = num_hash(column + 1)
            cell_name = "{}{}".format(cell1Column, row)
            cell_name2 = "{}{}".format(cell2Column, row)
            cell_name3 = "{}{}".format(cell3Column, row)
            if currentSheet[cell_name].value == Card_Name: #Can't manipulate the cell value so can't use upper(), title() etc...
                print("There are {} {} {}. It is at cell {}".format(currentSheet[cell_name3].value, currentSheet[cell_name].value, currentSheet[cell_name2].value ,cell_name))

def find_cardCode(currentSheet, Card_Code):
    for row in range(1, currentSheet.max_row + 1):
        for column in range(1,703): #columns
            cell1Column = num_hash(column)
            cell2Column = num_hash(column + 1)
            cell3Column = num_hash(column + 2)
            cell_name = "{}{}".format(cell1Column, row)
            cell_name2 = "{}{}".format(cell2Column, row)
            cell_name3 = "{}{}".format(cell3Column, row)
            if currentSheet[cell_name].value == Card_Code: #Can't manipulate the cell value so can't use upper(), title() etc...
                print("There are {} {} {}. It is at cell {}".format(currentSheet[cell_name3].value, currentSheet[cell_name].value, currentSheet[cell_name2].value ,cell_name))

def num_hash(num):
    alpha = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    if num < 26:
        return alpha[num-1]
    else:
        quotient, remainder = num // 26, num % 26
        if remainder == 0:
            if quotient == 1:
                return alpha[remainder-1]
            else:
                return num_hash(quotient-1) + alpha[remainder-1]
        else:
            return num_hash(quotient) + alpha[remainder-1]

def find_cardBYname():
    Card_Name = input("Enter the name of the card you wish to find: ").title() # Search result is Case sensitive

    for sheet in SheetNames:
        print("\nCurrent sheet is: {}".format(sheet) + "\n")
        currentSheet = File[sheet]
        find_cardName(currentSheet, Card_Name)

def find_cardBYcode():
    Card_Code = input("Enter the name of the card you wish to find: ").title() # Search result is Case sensitive

    for sheet in SheetNames:
        print("\nCurrent sheet is: {}".format(sheet) + "\n")
        currentSheet = File[sheet]
        find_cardCode(currentSheet, Card_Code)

def inputChoice():
    codeORname = input("Press \"C\" if you want to find card by its code or \"N\" if you want to search by its name: ").upper()
    if (codeORname == "C"):
        find_cardBYcode()
    if (codeORname == "N"):
        find_cardBYname()
    else:
        print("Not valid input")

#Variables / Entry
relativePath = os.path.join(RootPath, "FFTCG.xlsx")
File = openpyxl.load_workbook(relativePath, data_only=True)
SheetNames = File.sheetnames
print("Welcome to the FFTCG spreadsheet searcher!\n")
inputChoice()

choice = True
while True:
    choice = input("Do you want to find another card? (Y = yes, N = no)\n").upper()
    if (choice == "Y"):
        choice == True
        inputChoice()

    if (choice == "N"):
        choice == False
        print("Thank you for using this Python script!")
        break