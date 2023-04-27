from ColumnConverter import num_hash
from path import File, Path
import re
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill
from openpyxl.formatting.rule import CellIsRule, Rule
import time

def getInput(msg):
    return input(msg).upper()

def find_cardLocation(currentSheet, Card, searchType):
    s_time = time.time()
    cardNameRegex = re.match('^\d{1,2}-\d{3}[CRHLS]+$', Card)
    for row in range(1, currentSheet.max_row + 1):
        for column in range(1,currentSheet.max_column + 1): #columns
            startColumn = num_hash(column)
            leftColumn = num_hash(column - 1)
            rightColumn = num_hash(column + 1)
            secondRightColumn = num_hash(column + 2)
            searchTarget = f"{startColumn}{row}"
            pile = f"{startColumn}2"
            left_cell = f"{leftColumn}{row}"
            right_cell = f"{rightColumn}{row}"
            right_2Cells = f"{secondRightColumn}{row}"

            if (currentSheet[searchTarget].value == Card and cardNameRegex is None and searchType == "name"): #Can't manipulate the cell value so can't use upper(), title() etc...
                print(f"There are {currentSheet[right_cell].value} {currentSheet[left_cell].value} {currentSheet[searchTarget].value}(s). It is in the {currentSheet[pile].value} pile at Cell {searchTarget}.")
                print(f"\nTime taken: {time.time() - s_time} seconds.\n")

            if (currentSheet[searchTarget].value == Card and cardNameRegex is not None and searchType == "code"): #re.match returns None if no matches are found
                print(f"There are {currentSheet[right_2Cells].value} {currentSheet[searchTarget].value} {currentSheet[right_cell].value}(s) at Cell {searchTarget}.")
                print(f"\nTime taken: {time.time() - s_time} seconds.\n")

def emptyCells(ActiveWorksheet):
    print("Checking for empty cells in this sheet...\n")
    for row in range(1, ActiveWorksheet.max_row + 1):
        for column in range(1, ActiveWorksheet.max_column + 1):
            startColumn = num_hash(column)
            top = f"{num_hash(column)}1"
            searchTarget = f"{startColumn}{row}"
            if (ActiveWorksheet[searchTarget].value is None and ActiveWorksheet[top].value == "Code"):
                print(searchTarget)

def fillEmptyCell(ActiveWorksheet):
    print("Reminder, empty cells displayed are under the \"Code\" column!")
    # Create NamedStyle for reusability
    cell_format = NamedStyle(name = 'apply_format')
    cell_format.font = Font(bold=True)
    cell_format.alignment = Alignment(horizontal='center')
    # Input location of cell and get the 2 cells next to them
    selected_cell = getInput("Choose the cell you want to write to:\n")
    selectedCells = [selected_cell, ActiveWorksheet[selected_cell].offset(row=0, column=1).coordinate,
                     ActiveWorksheet[selected_cell].offset(row=0, column=2).coordinate]
    # Input value for cell
    change_value = [getInput(f"Enter what you want to put in cell {cell}:\n") for cell in selectedCells]
    #Assign new value
    for cell, value in zip(selectedCells, change_value):
        ActiveWorksheet[cell].value = value
    #Add checker so that it adds style to workbook if it isn't avaliable (like for instance on another PC)
    try:
        File.add_named_style(cell_format)
    except ValueError:
        pass
    #Apply style to cell after knowing where the cell is
    for cell in selectedCells:
        ActiveWorksheet[cell].style = 'apply_format'
    # Sucessful output
    print(f"\"{ActiveWorksheet[selectedCells[0]].value}\", \"{ActiveWorksheet[selectedCells[1]].value}\", \"{ActiveWorksheet[selectedCells[2]].value}\" has successfully been inputted into the spreadsheet.")
    # Save file
    File.save(Path)
    print("File saved.")

def reapply_conditional_formatting(ActiveWorksheet): ##Reapply conditional formatting rules after deleting values
    # TODO: Get current active sheet then reapply conditional formatting to the "Code" column.
    #Color
    duplicateColor = CellIsRule(operator='equal', formula=['1'], fill=PatternFill(bgColor="FFFF00"))
    print(f"Current sheet is: {ActiveWorksheet}\n")
    for rule in ActiveWorksheet.conditional_formatting._cf_rules:
        print(list(rule))
        rule = Rule(type='DuplicateValues', dxf=duplicateColor, stopIfTrue=True, formula=['1'])
        # print(rule.sqref)
        # if top cell is equal to "Code" then apply conditional formatting. Apply the conditional format to the whole column.
        for column in range(1, ActiveWorksheet.max_column + 1):
            code_cell = f"{num_hash(column)}1"
            if (ActiveWorksheet[code_cell].value == "Code"):
                print(code_cell)
                letter = code_cell.column_letter
                ActiveWorksheet.conditional_formatting.add(f"{letter}:{letter}", rule)
                # print(f"Conditional formatting has been reapplied to the \"Code\" column of {ActiveWorksheet}.")
                # break